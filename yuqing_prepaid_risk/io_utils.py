from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .config import BASE_COLUMNS, RISK_COLUMNS
from .models import ProcessedItem

def normalize_row(item: Dict[str, Any]) -> Dict[str, Any]:
    row = {col: item.get(col, "") for col in BASE_COLUMNS}
    row["iplocation"] = item.get("iplocation") or item.get("ipLocation") or ""
    row["images"] = normalize_images(row.get("images"))
    return row


def normalize_images(images: Any) -> str:
    if images is None:
        return ""
    if isinstance(images, str):
        return images
    if isinstance(images, list):
        return "[" + ", ".join(str(x) for x in images) + "]"
    return str(images)


def read_local(path: Path) -> List[Dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return read_xlsx(path)
    if suffix == ".json":
        data = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(data, dict):
            data = data.get("data_list") or data.get("data") or []
        return [normalize_row(x) for x in data]
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            return [normalize_row(row) for row in csv.DictReader(f)]
    raise ValueError(f"不支持的输入文件格式: {path}")


def read_xlsx(path: Path) -> List[Dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    headers = [str(x) if x is not None else "" for x in next(rows_iter)]
    rows: List[Dict[str, Any]] = []
    for values in rows_iter:
        item = dict(zip(headers, values))
        rows.append(normalize_row(item))
    return rows

def write_xlsx(items: List[ProcessedItem], output: Path, include_filtered: bool) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Result"
    columns = BASE_COLUMNS + RISK_COLUMNS
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="366092")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for item in items:
        if not include_filtered and not item.keep:
            continue
        row = dict(item.row)
        row.update(
            {
                "risk_level": item.risk_level,
                "risk_label": item.risk_label,
                "risk_summary": item.risk_summary,
                "dedup_key": item.dedup_key,
                "dedup_reason": item.dedup_reason,
                "model_used": item.model_used,
                "model_reason": item.model_reason,
            }
        )
        ws.append([row.get(col, "") for col in columns])

    widths = {
        "A": 8,
        "B": 42,
        "C": 46,
        "D": 20,
        "E": 70,
        "F": 12,
        "G": 12,
        "H": 16,
        "I": 20,
        "J": 30,
        "K": 14,
        "L": 22,
        "M": 18,
        "N": 80,
        "O": 28,
        "P": 24,
        "Q": 12,
        "R": 42,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
