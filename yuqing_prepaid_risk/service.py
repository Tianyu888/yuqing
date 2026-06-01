from __future__ import annotations

import os
import time
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .api import fetch_yuqing, post_api
from .config import (
    APP_KEY,
    BASE_COLUMNS,
    DEFAULT_LLM_API_KEY,
    DEFAULT_LLM_API_URL,
    DEFAULT_LLM_MODEL,
    DEFAULT_THEME_NAME,
    SECURE_KEY,
    THEME_API_URL,
)
from .env import load_env_file
from .io_utils import read_local, write_xlsx
from .llm import normalize_chat_completions_url
from .models import ProcessedItem
from .pipeline import process_rows
from .utils import log

DEFAULT_DISTRICTS = "梁溪区,锡山区,惠山区,滨湖区,新吴区,江阴市,宜兴市,经开区,无锡"
GPU_LLM_REQUIRED_MESSAGE = "analysisMode=gpu 模式下需要配置大模型，请查阅文档进行大模型API的相关配置"


@dataclass
class RiskAnalysisOptions:
    input: Optional[Path | str] = None
    output: Path | str = Path("预充值卡商户跑路风险预警输出.xlsx")
    include_filtered: bool = False
    include_suspected: bool = True
    app_key: Optional[str] = None
    secure_key: Optional[str] = None
    theme_name: str = DEFAULT_THEME_NAME
    news_type: str = "all"
    attribute: str = "negative"
    date_type: int = 7
    begin_date: Optional[str] = None
    end_date: Optional[str] = None
    page_count: int = 20
    max_pages: int = 200
    timeout: int = 20
    sleep: float = 0.2
    payload_format: str = "json"
    llm_api_url: Optional[str] = None
    llm_api_key: Optional[str] = None
    llm_model: Optional[str] = None
    llm_timeout: int = 120
    llm_content_chars: int = 5000
    analysis_mode: str = "cpu"
    llm_workers: int = 4
    disable_llm: bool = False
    target_city: str = "无锡市"
    target_province: str = "江苏省"
    districts: str | Sequence[str] = DEFAULT_DISTRICTS
    fuzzy_threshold: float = 0.90


@dataclass
class RawExportOptions:
    output: Path | str = Path("outputs/舆情API原始结果.xlsx")
    app_key: Optional[str] = None
    secure_key: Optional[str] = None
    url: str = THEME_API_URL
    theme_name: str = DEFAULT_THEME_NAME
    news_type: str = "all"
    attribute: str = "negative"
    date_type: int = 7
    begin_date: Optional[str] = None
    end_date: Optional[str] = None
    page_count: int = 20
    max_pages: int = 200
    timeout: int = 20
    sleep: float = 0.2
    payload_format: str = "json"


@dataclass
class YuqingApiTestOptions:
    app_key: Optional[str] = None
    secure_key: Optional[str] = None
    url: str = THEME_API_URL
    theme_name: str = DEFAULT_THEME_NAME
    timeout: int = 20
    payload_format: str = "json"


@dataclass
class LlmApiTestOptions:
    model_url: Optional[str] = None
    model_key: Optional[str] = None
    model_name: Optional[str] = None
    timeout: int = 60


def _load_env() -> Dict[str, str]:
    return load_env_file()


def _namespace(values: Dict[str, Any]) -> Any:
    class OptionsNamespace:
        pass

    obj = OptionsNamespace()
    for key, value in values.items():
        setattr(obj, key, value)
    return obj


def _district_list(value: str | Sequence[str]) -> List[str]:
    if isinstance(value, str):
        return [x.strip() for x in value.split(",") if x.strip()]
    return [str(x).strip() for x in value if str(x).strip()]


def _risk_namespace(options: RiskAnalysisOptions) -> Any:
    env = _load_env()
    values = asdict(options)
    values["input"] = Path(options.input) if options.input else None
    values["output"] = Path(options.output)
    values["app_key"] = options.app_key or os.getenv("SUYIYU_APP_KEY") or env.get("SUYIYU_APP_KEY", APP_KEY)
    values["secure_key"] = options.secure_key or os.getenv("SUYIYU_SECURE_KEY") or env.get("SUYIYU_SECURE_KEY", SECURE_KEY)
    values["llm_api_url"] = options.llm_api_url or os.getenv("LLM_API_URL") or env.get("LLM_API_URL", DEFAULT_LLM_API_URL)
    values["llm_api_key"] = options.llm_api_key or os.getenv("LLM_API_KEY") or env.get("LLM_API_KEY", DEFAULT_LLM_API_KEY)
    values["llm_model"] = options.llm_model or os.getenv("LLM_MODEL") or env.get("LLM_MODEL", DEFAULT_LLM_MODEL)
    return _namespace(values)


def _raw_namespace(options: RawExportOptions) -> Any:
    env = _load_env()
    values = asdict(options)
    values["output"] = Path(options.output)
    values["app_key"] = options.app_key or os.getenv("SUYIYU_APP_KEY") or env.get("SUYIYU_APP_KEY", APP_KEY)
    values["secure_key"] = options.secure_key or os.getenv("SUYIYU_SECURE_KEY") or env.get("SUYIYU_SECURE_KEY", SECURE_KEY)
    return _namespace(values)


def validate_risk_options(args: Any) -> None:
    if str(args.analysis_mode).lower() != "gpu" or args.disable_llm:
        return
    if args.llm_api_url and args.llm_api_key and args.llm_model:
        return
    raise ValueError(GPU_LLM_REQUIRED_MESSAGE)


def processed_item_to_dict(item: ProcessedItem) -> Dict[str, Any]:
    row = dict(item.row)
    row.update(
        {
            "risk_level": item.risk_level,
            "risk_label": item.risk_label,
            "risk_summary": item.risk_summary,
            "dedup_key": item.dedup_key,
            "dedup_reason": item.dedup_reason,
            "keep": item.keep,
            "filter_reason": item.filter_reason,
            "model_used": item.model_used,
            "model_reason": item.model_reason,
        }
    )
    return row


def run_risk_analysis(
    options: Optional[RiskAnalysisOptions | Dict[str, Any]] = None,
    *,
    write_output: bool = True,
    include_items: bool = True,
) -> Dict[str, Any]:
    """Run the full prepaid-risk workflow as an importable service API."""
    opts = RiskAnalysisOptions(**options) if isinstance(options, dict) else (options or RiskAnalysisOptions())
    args = _risk_namespace(opts)
    validate_risk_options(args)
    districts = _district_list(args.districts)

    if args.input:
        rows = read_local(args.input)
        log(f"本地舆情读取完成: input={args.input}, rows={len(rows)}")
    else:
        rows = fetch_yuqing(args)

    items, stats = process_rows(
        rows,
        args.target_city,
        args.target_province,
        districts,
        args.fuzzy_threshold,
        args.include_suspected,
        args,
    )
    if write_output:
        write_xlsx(items, args.output, args.include_filtered)

    result: Dict[str, Any] = {
        "stats": stats,
        "output": str(args.output),
        "input_count": len(rows),
        "kept_count": stats.get("输出总量", 0),
    }
    if include_items:
        result["items"] = [processed_item_to_dict(item) for item in items]
    return result


def raw_cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    import json

    return json.dumps(value, ensure_ascii=False)


def build_raw_columns(rows: List[Dict[str, Any]]) -> List[str]:
    seen = set()
    columns: List[str] = []
    for key in BASE_COLUMNS:
        if key not in seen:
            columns.append(key)
            seen.add(key)
    for row in rows:
        for key in row:
            if key not in seen:
                columns.append(str(key))
                seen.add(key)
    return columns


def write_raw_xlsx(rows: List[Dict[str, Any]], output: Path | str) -> None:
    output_path = Path(output)
    wb = Workbook()
    ws = wb.active
    ws.title = "RawData"
    columns = build_raw_columns(rows)
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="366092")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        ws.append([raw_cell_value(row.get(column, "")) for column in columns])

    for idx, column in enumerate(columns, start=1):
        width = 18
        if column in {"title", "content"}:
            width = 60 if column == "title" else 90
        elif column in {"url", "images"}:
            width = 50
        elif column in {"pubtime", "source", "medianame"}:
            width = 22
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def fetch_raw_yuqing_rows(options: Optional[RawExportOptions | Dict[str, Any]] = None) -> List[Dict[str, Any]]:
    opts = RawExportOptions(**options) if isinstance(options, dict) else (options or RawExportOptions())
    args = _raw_namespace(opts)
    if not args.app_key or not args.secure_key:
        raise ValueError("缺少 SUYIYU_APP_KEY 或 SUYIYU_SECURE_KEY，请先配置 .env 或传入 app_key/secure_key")

    base_params: Dict[str, Any] = {
        "theme_name": args.theme_name,
        "news_type": args.news_type,
        "attribute": args.attribute,
        "date_type": args.date_type,
    }
    if args.date_type == 0:
        base_params["begin_date"] = args.begin_date
        base_params["end_date"] = args.end_date

    rows: List[Dict[str, Any]] = []
    for page in range(1, args.max_pages + 1):
        params = dict(base_params)
        params["page"] = page
        params["page_count"] = args.page_count
        payload = post_api(args.url, params, args.app_key, args.secure_key, args.timeout, args.payload_format)
        data = payload.get("data") or {}
        data_list = data.get("data_list") or []
        rows.extend(item for item in data_list if isinstance(item, dict))
        total_num = int(data.get("total_num") or 0)
        total_pages = (total_num + args.page_count - 1) // args.page_count if args.page_count else 0
        log(f"已拉取第 {page}/{total_pages or '?'} 页: 本页={len(data_list)}, 累计={len(rows)}, total={total_num}")
        if not data_list or len(rows) >= total_num:
            break
        time.sleep(args.sleep)
    return rows


def export_raw_yuqing(
    options: Optional[RawExportOptions | Dict[str, Any]] = None,
    *,
    write_output: bool = True,
) -> Dict[str, Any]:
    opts = RawExportOptions(**options) if isinstance(options, dict) else (options or RawExportOptions())
    args = _raw_namespace(opts)
    rows = fetch_raw_yuqing_rows(opts)
    if write_output:
        write_raw_xlsx(rows, args.output)
    return {"rows": rows, "row_count": len(rows), "output": str(args.output)}


def test_yuqing_api(options: Optional[YuqingApiTestOptions | Dict[str, Any]] = None) -> Dict[str, Any]:
    opts = YuqingApiTestOptions(**options) if isinstance(options, dict) else (options or YuqingApiTestOptions())
    env = _load_env()
    app_key = opts.app_key or os.getenv("SUYIYU_APP_KEY") or env.get("SUYIYU_APP_KEY", APP_KEY)
    secure_key = opts.secure_key or os.getenv("SUYIYU_SECURE_KEY") or env.get("SUYIYU_SECURE_KEY", SECURE_KEY)
    if not app_key or not secure_key:
        raise ValueError("缺少 SUYIYU_APP_KEY 或 SUYIYU_SECURE_KEY")
    params = {
        "theme_name": opts.theme_name,
        "news_type": "all",
        "attribute": "negative",
        "date_type": 1,
        "page": 1,
        "page_count": 1,
    }
    payload = post_api(opts.url, params, app_key, secure_key, opts.timeout, opts.payload_format)
    return {"ok": True, "payload": payload, "fields": list(payload)[:10]}


def test_llm_api(options: Optional[LlmApiTestOptions | Dict[str, Any]] = None) -> Dict[str, Any]:
    opts = LlmApiTestOptions(**options) if isinstance(options, dict) else (options or LlmApiTestOptions())
    env = _load_env()
    model_url = opts.model_url or os.getenv("LLM_API_URL") or env.get("LLM_API_URL", DEFAULT_LLM_API_URL)
    model_key = opts.model_key or os.getenv("LLM_API_KEY") or env.get("LLM_API_KEY", DEFAULT_LLM_API_KEY)
    model_name = opts.model_name or os.getenv("LLM_MODEL") or env.get("LLM_MODEL", DEFAULT_LLM_MODEL)
    if not model_url or not model_key or not model_name:
        raise ValueError("缺少 LLM_API_URL、LLM_API_KEY 或 LLM_MODEL")

    payload = {"model": model_name, "messages": [{"role": "user", "content": "请只回复 OK"}]}
    headers = {"Authorization": f"Bearer {model_key}", "Content-Type": "application/json"}
    request_url = normalize_chat_completions_url(model_url)
    response = requests.post(request_url, json=payload, headers=headers, timeout=opts.timeout)
    response.raise_for_status()
    data = response.json()
    content = data["choices"][0]["message"]["content"]
    return {"ok": True, "content": content, "request_url": request_url, "payload": data}
