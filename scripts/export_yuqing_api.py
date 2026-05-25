#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import json
import os
import sys
import time
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from yuqing_prepaid_risk.api import post_api
from yuqing_prepaid_risk.config import BASE_COLUMNS, DEFAULT_THEME_NAME, THEME_API_URL
from yuqing_prepaid_risk.env import load_env_file
from yuqing_prepaid_risk.utils import log


def cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return json.dumps(value, ensure_ascii=False)


def build_columns(rows: List[Dict[str, Any]]) -> List[str]:
    seen = set()
    columns: List[str] = []
    for key in BASE_COLUMNS:
        if key not in seen:
            columns.append(key)
            seen.add(key)
    for row in rows:
        for key in row:
            if key not in seen:
                columns.append(str(key))
                seen.add(key)
    return columns


def write_raw_xlsx(rows: List[Dict[str, Any]], output: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "RawData"
    columns = build_columns(rows)
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="366092")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        ws.append([cell_value(row.get(column, "")) for column in columns])

    for idx, column in enumerate(columns, start=1):
        width = 18
        if column in {"title", "content"}:
            width = 60 if column == "title" else 90
        elif column in {"url", "images"}:
            width = 50
        elif column in {"pubtime", "source", "medianame"}:
            width = 22
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)


def parse_args() -> argparse.Namespace:
    load_env_file()
    parser = argparse.ArgumentParser(description="直接拉取苏移舆情 API 原始结果并导出 Excel")
    parser.add_argument("--output", type=Path, default=Path("outputs/舆情API原始结果.xlsx"))
    parser.add_argument("--app-key", default=os.getenv("SUYIYU_APP_KEY", ""))
    parser.add_argument("--secure-key", default=os.getenv("SUYIYU_SECURE_KEY", ""))
    parser.add_argument("--url", default=THEME_API_URL)
    parser.add_argument("--theme-name", default=DEFAULT_THEME_NAME)
    parser.add_argument("--news-type", default="all")
    parser.add_argument("--attribute", default="negative", choices=["all", "positive", "neutral", "negative"])
    parser.add_argument("--date-type", type=int, default=7, choices=[0, 1, 7, 30])
    parser.add_argument("--begin-date")
    parser.add_argument("--end-date")
    parser.add_argument("--page-count", type=int, default=20)
    parser.add_argument("--max-pages", type=int, default=200)
    parser.add_argument("--timeout", type=int, default=20)
    parser.add_argument("--sleep", type=float, default=0.2)
    parser.add_argument("--payload-format", choices=["json", "form"], default="json")
    return parser.parse_args()


def fetch_raw_rows(args: argparse.Namespace) -> List[Dict[str, Any]]:
    if not args.app_key or not args.secure_key:
        raise ValueError("缺少 SUYIYU_APP_KEY 或 SUYIYU_SECURE_KEY，请先配置 .env 或使用 --app-key/--secure-key")

    base_params: Dict[str, Any] = {
        "theme_name": args.theme_name,
        "news_type": args.news_type,
        "attribute": args.attribute,
        "date_type": args.date_type,
    }
    if args.date_type == 0:
        base_params["begin_date"] = args.begin_date
        base_params["end_date"] = args.end_date

    rows: List[Dict[str, Any]] = []
    for page in range(1, args.max_pages + 1):
        params = dict(base_params)
        params["page"] = page
        params["page_count"] = args.page_count
        payload = post_api(args.url, params, args.app_key, args.secure_key, args.timeout, args.payload_format)
        data = payload.get("data") or {}
        data_list = data.get("data_list") or []
        rows.extend(item for item in data_list if isinstance(item, dict))
        total_num = int(data.get("total_num") or 0)
        log(f"已拉取第 {page} 页: 本页={len(data_list)}, 累计={len(rows)}, total={total_num}")
        if not data_list or len(rows) >= total_num:
            break
        time.sleep(args.sleep)
    return rows


def main() -> int:
    args = parse_args()
    rows = fetch_raw_rows(args)
    write_raw_xlsx(rows, args.output)
    print(f"舆情 API 原始结果导出完成: rows={len(rows)}")
    print(f"输出文件: {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
